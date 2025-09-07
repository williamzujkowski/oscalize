"""
Inventory reader for FedRAMP Integrated Inventory Workbook (Attachment 13)

Reads Excel inventory templates and converts to CIR format.
Enforces FedRAMP IIW structure and provides detailed error reporting.
"""

import logging
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook

from .base_reader import BaseReader

logger = logging.getLogger(__name__)


class InventoryReader(BaseReader):
    """Reader for FedRAMP Integrated Inventory Workbook Excel files"""
    
    # Expected column mappings for FedRAMP IIW
    EXPECTED_COLUMNS = {
        'Asset ID': 'asset_id',
        'Asset Type': 'asset_type', 
        'Asset Name': 'name',
        'Asset Description': 'description',
        'Environment': 'environment',
        'Service Layer': 'service_layer',
        'Function': 'function',
        'Public (Internet Accessible)': 'public_access',
        'Virtual (Y/N)': 'virtual',
        'IP Address': 'ip_address',
        'MAC Address': 'mac_address', 
        'VLAN': 'vlan',
        'Network Location': 'network_location',
        'Asset Owner': 'asset_owner',
        'System Administrator': 'system_admin',
        'Data Sensitivity/Criticality': 'criticality',
        'Baseline Configuration': 'baseline',
        'Operating System': 'operating_system',
        'Software/Application Version': 'software_version',
        'Patch Level': 'patch_level'
    }
    
    ASSET_TYPES = ['hardware', 'software', 'data', 'network', 'service', 'other']
    ENVIRONMENTS = ['Production', 'Development', 'Test', 'Staging', 'Other']
    CRITICALITY_LEVELS = ['Low', 'Moderate', 'High', 'Critical']
    
    def __init__(self, file_path: Path):
        super().__init__(file_path)
        self.workbook = None
        self.sheet_name = None
        
    def to_cir(self) -> Dict[str, Any]:
        """Convert Inventory Excel file to CIR format"""
        logger.info(f"Converting Inventory Excel file: {self.file_path}")
        
        # Load workbook to detect sheet structure
        self.workbook = load_workbook(self.file_path, data_only=True)
        self.sheet_name = self._find_inventory_sheet()
        
        if not self.sheet_name:
            raise ValueError("No inventory sheet found in workbook")
        
        # Read data using pandas
        df = pd.read_excel(self.file_path, sheet_name=self.sheet_name)
        
        # Detect template version
        template_version = self._detect_template_version(df)
        
        # Validate and normalize column structure
        column_mapping = self._validate_columns(df)
        
        # Process assets
        assets = self._process_assets(df, column_mapping)
        
        return {
            "metadata": self._create_base_metadata(
                "xlsx",
                sheet_name=self.sheet_name,
                template_version=template_version
            ),
            "assets": assets
        }
    
    def _find_inventory_sheet(self) -> Optional[str]:
        """Find the inventory sheet in the workbook"""
        sheet_names = self.workbook.sheetnames
        
        # Look for sheets with inventory-related names
        inventory_keywords = ['inventory', 'asset', 'component', 'system']
        
        for sheet_name in sheet_names:
            sheet_lower = sheet_name.lower()
            for keyword in inventory_keywords:
                if keyword in sheet_lower:
                    return sheet_name
        
        # Fallback to first sheet
        if sheet_names:
            logger.warning(f"No inventory sheet found, using first sheet: {sheet_names[0]}")
            return sheet_names[0]
        
        return None
    
    def _detect_template_version(self, df: pd.DataFrame) -> str:
        """Detect FedRAMP IIW template version"""
        columns = set(col.strip().lower() for col in df.columns if pd.notna(col))
        
        # Check for FedRAMP IIW specific columns
        iiw_indicators = {'asset id', 'asset type', 'asset name', 'environment'}
        
        if iiw_indicators.issubset(columns):
            return "FedRAMP_IIW_v4.0"
        
        logger.warning("Could not detect IIW template version, assuming compatible format")
        return "Compatible_IIW_format"
    
    def _validate_columns(self, df: pd.DataFrame) -> Dict[str, str]:
        """Validate and create column mapping"""
        available_columns = set(df.columns.str.strip())
        column_mapping = {}
        missing_required = []
        
        # Map available columns to expected fields
        for expected_col, field_name in self.EXPECTED_COLUMNS.items():
            # Try exact match first
            if expected_col in available_columns:
                column_mapping[expected_col] = field_name
                continue
            
            # Try case-insensitive and fuzzy matching
            matched = False
            for available_col in available_columns:
                if self._fuzzy_column_match(expected_col, available_col):
                    column_mapping[available_col] = field_name
                    matched = True
                    break
            
            if not matched:
                # Check if this is a required field
                if field_name in ['asset_id', 'asset_type', 'name', 'environment', 'criticality']:
                    missing_required.append(expected_col)
                else:
                    logger.warning(f"Optional column not found: {expected_col}")
        
        if missing_required:
            raise ValueError(
                f"Missing required inventory columns: {', '.join(missing_required)}\n"
                f"Available columns: {', '.join(sorted(available_columns))}"
            )
        
        return {col: field for col, field in column_mapping.items()}
    
    def _fuzzy_column_match(self, expected: str, available: str) -> bool:
        """Fuzzy matching for column names"""
        expected_clean = expected.lower().replace(' ', '').replace('/', '').replace('(', '').replace(')', '')
        available_clean = available.lower().replace(' ', '').replace('/', '').replace('(', '').replace(')', '')
        
        # Exact match after cleaning
        if expected_clean == available_clean:
            return True
        
        # Check if key terms are present
        expected_terms = set(expected.lower().split())
        available_terms = set(available.lower().split())
        
        # Require at least 2 matching terms or 1 term if it's unique
        common_terms = expected_terms.intersection(available_terms)
        if len(common_terms) >= 2 or (len(common_terms) == 1 and len(expected_terms) == 1):
            return True
        
        return False
    
    def _process_assets(self, df: pd.DataFrame, column_mapping: Dict[str, str]) -> List[Dict[str, Any]]:
        """Process inventory assets and convert to CIR format"""
        assets = []
        
        for idx, row in df.iterrows():
            # Skip empty rows
            if pd.isna(row.get(list(column_mapping.keys())[0])):
                continue
            
            asset_data = {}
            row_num = idx + 2  # Excel is 1-indexed, plus header row
            
            # Process each mapped column
            for excel_col, field_name in column_mapping.items():
                value = row.get(excel_col)
                processed_value = self._process_field_value(field_name, value, row_num, excel_col)
                asset_data[field_name] = processed_value
            
            # Add computed fields
            asset_data['tags'] = self._extract_tags(asset_data)
            asset_data['links'] = self._extract_links(asset_data)
            
            # Add source attribution
            asset_data['source'] = self._create_source_reference(
                sheet=self.sheet_name,
                row=row_num,
                col_range=f"A{row_num}:{self._get_last_column_letter(len(column_mapping))}{row_num}"
            )
            
            # Validate required fields
            self._validate_asset_data(asset_data, row_num)
            
            assets.append(asset_data)
        
        logger.info(f"Processed {len(assets)} inventory assets")
        return assets
    
    def _process_field_value(self, field_name: str, value: Any, row_num: int, excel_col: str) -> Any:
        """Process individual field values with type conversion and validation"""
        if pd.isna(value):
            return ""
        
        # Convert to string and strip whitespace
        str_value = str(value).strip()
        
        if field_name == 'asset_type':
            # Normalize asset type
            normalized = self._normalize_asset_type(str_value)
            if normalized not in self.ASSET_TYPES:
                logger.warning(f"Unknown asset type '{str_value}' at row {row_num}, column {excel_col}")
            return normalized
        
        elif field_name == 'environment':
            # Validate environment
            if str_value not in self.ENVIRONMENTS:
                logger.warning(f"Unknown environment '{str_value}' at row {row_num}, column {excel_col}")
                # Try to normalize
                str_value = self._normalize_environment(str_value)
            return str_value
        
        elif field_name == 'criticality':
            # Validate criticality
            normalized = self._normalize_criticality(str_value)
            if normalized not in self.CRITICALITY_LEVELS:
                logger.warning(f"Unknown criticality '{str_value}' at row {row_num}, column {excel_col}")
            return normalized
        
        elif field_name in ['public_access', 'virtual']:
            # Convert to boolean
            return self._parse_boolean(str_value)
        
        elif field_name == 'ip_address':
            # Basic IP validation
            if str_value and not self._is_valid_ip_format(str_value):
                logger.warning(f"Invalid IP format '{str_value}' at row {row_num}, column {excel_col}")
            return str_value
        
        else:
            return str_value if str_value else ""
    
    def _normalize_asset_type(self, value: str) -> str:
        """Normalize asset type values"""
        value_lower = value.lower()
        
        if 'hardware' in value_lower or 'hw' in value_lower:
            return 'hardware'
        elif 'software' in value_lower or 'sw' in value_lower or 'application' in value_lower:
            return 'software'
        elif 'data' in value_lower or 'database' in value_lower:
            return 'data'
        elif 'network' in value_lower or 'net' in value_lower:
            return 'network'
        elif 'service' in value_lower or 'svc' in value_lower:
            return 'service'
        else:
            return 'other'
    
    def _normalize_environment(self, value: str) -> str:
        """Normalize environment values"""
        value_lower = value.lower()
        
        if value_lower in ['prod', 'production']:
            return 'Production'
        elif value_lower in ['dev', 'development']:
            return 'Development'
        elif value_lower in ['test', 'testing', 'qa']:
            return 'Test'
        elif value_lower in ['stage', 'staging']:
            return 'Staging'
        
        return value  # Return original if no match
    
    def _normalize_criticality(self, value: str) -> str:
        """Normalize criticality values"""
        value_lower = value.lower()
        
        if value_lower in ['low', 'l']:
            return 'Low'
        elif value_lower in ['moderate', 'med', 'medium', 'm']:
            return 'Moderate'
        elif value_lower in ['high', 'h']:
            return 'High'
        elif value_lower in ['critical', 'crit', 'c']:
            return 'Critical'
        
        return value  # Return original if no match
    
    def _parse_boolean(self, value: str) -> bool:
        """Parse boolean values from various formats"""
        value_lower = value.lower()
        
        if value_lower in ['yes', 'y', 'true', '1', 'on']:
            return True
        elif value_lower in ['no', 'n', 'false', '0', 'off']:
            return False
        
        # Default to False for unclear values
        return False
    
    def _is_valid_ip_format(self, ip_str: str) -> bool:
        """Basic IP address format validation"""
        import re
        
        # IPv4 pattern
        ipv4_pattern = r'^(\d{1,3}\.){3}\d{1,3}$'
        # IPv6 pattern (simplified)
        ipv6_pattern = r'^([0-9a-fA-F]{1,4}:){7}[0-9a-fA-F]{1,4}$'
        
        return bool(re.match(ipv4_pattern, ip_str) or re.match(ipv6_pattern, ip_str))
    
    def _extract_tags(self, asset_data: Dict[str, Any]) -> List[str]:
        """Extract tags from asset data"""
        tags = []
        
        # Add environment as tag
        if asset_data.get('environment'):
            tags.append(f"env:{asset_data['environment'].lower()}")
        
        # Add criticality as tag
        if asset_data.get('criticality'):
            tags.append(f"criticality:{asset_data['criticality'].lower()}")
        
        # Add asset type as tag
        if asset_data.get('asset_type'):
            tags.append(f"type:{asset_data['asset_type']}")
        
        # Add special tags
        if asset_data.get('public_access'):
            tags.append('public-facing')
        
        if asset_data.get('virtual'):
            tags.append('virtual')
        
        return tags
    
    def _extract_links(self, asset_data: Dict[str, Any]) -> List[Dict[str, str]]:
        """Extract links from asset data"""
        links = []
        
        # Add baseline configuration link if available
        if asset_data.get('baseline'):
            links.append({
                'rel': 'baseline-configuration',
                'href': f"#baseline-{asset_data.get('asset_id', 'unknown')}",
                'media_type': 'text/plain'
            })
        
        return links
    
    def _validate_asset_data(self, asset_data: Dict[str, Any], row_num: int) -> None:
        """Validate required fields in asset data"""
        required_fields = ['asset_id', 'asset_type', 'name', 'environment', 'criticality']
        
        for field in required_fields:
            if not asset_data.get(field):
                raise ValueError(f"Missing required field '{field}' at row {row_num}")
    
    def _get_last_column_letter(self, num_columns: int) -> str:
        """Convert column number to Excel column letter"""
        result = ""
        while num_columns > 0:
            num_columns -= 1
            result = chr(num_columns % 26 + ord('A')) + result
            num_columns //= 26
        return result