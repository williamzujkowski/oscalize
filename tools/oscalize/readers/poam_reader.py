"""
POA&M reader for Excel files

Reads FedRAMP POA&M v3.0 Excel templates and converts to CIR format.
Enforces template structure and provides detailed error reporting with cell coordinates.
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook

from .base_reader import BaseReader

logger = logging.getLogger(__name__)


class POAMReader(BaseReader):
    """Reader for FedRAMP POA&M v3.0 Excel files"""
    
    # Expected column mappings for FedRAMP POA&M v3.0
    EXPECTED_COLUMNS = {
        'POA&M Item ID': 'poam_id',
        'Vulnerability Description': 'title',
        'Security Control Number (NC/NH/NI)': 'control_ids',
        'Office/Organization': 'origin',
        'Security Control Name': 'control_name',
        'Implementation Guidance': 'implementation_guidance',
        'Severity': 'severity',
        'POA&M Status': 'status',
        'Scheduled Completion Date': 'scheduled_completion_date',
        'Actual Completion Date': 'actual_completion_date',
        'Point of Contact': 'point_of_contact',
        'Resources Required': 'resources_required',
        'Description': 'description',
        'Remediation Plan': 'remediation_plan',
        'Milestone Description': 'milestone_description',
        'Milestone Date': 'milestone_date',
        'Milestone Status': 'milestone_status',
        'Affected Assets': 'asset_ids',
        'Comments': 'comments'
    }
    
    SEVERITY_VALUES = ['Low', 'Moderate', 'High', 'Critical']
    STATUS_VALUES = ['Open', 'Ongoing', 'Completed', 'Risk Accepted']
    
    def __init__(self, file_path: Path):
        super().__init__(file_path)
        self.workbook = None
        self.sheet_name = None
        
    def to_cir(self) -> Dict[str, Any]:
        """Convert POA&M Excel file to CIR format"""
        logger.info(f"Converting POA&M Excel file: {self.file_path}")
        
        # Load workbook to detect sheet structure
        self.workbook = load_workbook(self.file_path, data_only=True)
        self.sheet_name = self._find_poam_sheet()
        
        if not self.sheet_name:
            raise ValueError("No POA&M sheet found in workbook")
        
        # Read data using pandas for easier processing
        df = pd.read_excel(self.file_path, sheet_name=self.sheet_name)
        
        # Detect template version
        template_version = self._detect_template_version(df)
        
        # Validate and normalize column structure
        column_mapping = self._validate_columns(df)
        
        # Process rows
        rows = self._process_rows(df, column_mapping)
        
        return {
            "metadata": self._create_base_metadata(
                "xlsx",
                sheet_name=self.sheet_name,
                template_version=template_version
            ),
            "rows": rows
        }
    
    def _find_poam_sheet(self) -> Optional[str]:
        """Find the POA&M sheet in the workbook"""
        sheet_names = self.workbook.sheetnames
        
        # Look for sheets with POA&M in the name
        for sheet_name in sheet_names:
            if 'poam' in sheet_name.lower() or 'poa&m' in sheet_name.lower():
                return sheet_name
        
        # Fallback to first sheet if no obvious POA&M sheet
        if sheet_names:
            logger.warning(f"No POA&M sheet found, using first sheet: {sheet_names[0]}")
            return sheet_names[0]
        
        return None
    
    def _detect_template_version(self, df: pd.DataFrame) -> str:
        """Detect FedRAMP POA&M template version"""
        columns = set(df.columns.str.strip().str.lower())
        
        # Check for v3.0 specific columns
        v30_indicators = {'poa&m item id', 'vulnerability description', 'security control number'}
        
        if v30_indicators.issubset({col.lower() for col in df.columns}):
            return "FedRAMP_POA&M_v3.0"
        
        logger.warning("Could not detect POA&M template version, assuming v3.0")
        return "Unknown_v3.0_compatible"
    
    def _validate_columns(self, df: pd.DataFrame) -> Dict[str, str]:
        """Validate and create column mapping"""
        available_columns = set(df.columns.str.strip())
        column_mapping = {}
        missing_required = []
        
        # Map available columns to expected fields
        for expected_col, field_name in self.EXPECTED_COLUMNS.items():
            # Try exact match first
            if expected_col in available_columns:
                column_mapping[expected_col] = field_name
                continue
            
            # Try case-insensitive match
            for available_col in available_columns:
                if expected_col.lower() == available_col.lower():
                    column_mapping[available_col] = field_name
                    break
            else:
                # Check if this is a required field
                if field_name in ['poam_id', 'title', 'severity', 'status']:
                    missing_required.append(expected_col)
                else:
                    logger.warning(f"Optional column not found: {expected_col}")
        
        if missing_required:
            raise ValueError(
                f"Missing required POA&M columns: {', '.join(missing_required)}\n"
                f"Available columns: {', '.join(sorted(available_columns))}"
            )
        
        return {col: field for col, field in column_mapping.items()}
    
    def _process_rows(self, df: pd.DataFrame, column_mapping: Dict[str, str]) -> List[Dict[str, Any]]:
        """Process POA&M rows and convert to CIR format"""
        rows = []
        
        for idx, row in df.iterrows():
            # Skip empty rows
            if pd.isna(row.get(list(column_mapping.keys())[0])):  # Check first mapped column
                continue
            
            row_data = {}
            row_num = idx + 2  # Excel is 1-indexed, plus header row
            
            # Process each mapped column
            for excel_col, field_name in column_mapping.items():
                value = row.get(excel_col)
                processed_value = self._process_field_value(field_name, value, row_num, excel_col)
                row_data[field_name] = processed_value
            
            # Add source attribution
            row_data['source'] = self._create_source_reference(
                sheet=self.sheet_name,
                row=row_num,
                col_range=f"A{row_num}:{self._get_last_column_letter(len(column_mapping))}{row_num}"
            )
            
            # Process milestones if present
            milestones = self._extract_milestones(row, row_num)
            if milestones:
                row_data['milestones'] = milestones
            
            # Validate required fields
            self._validate_row_data(row_data, row_num)
            
            rows.append(row_data)
        
        logger.info(f"Processed {len(rows)} POA&M items")
        return rows
    
    def _process_field_value(self, field_name: str, value: Any, row_num: int, excel_col: str) -> Any:
        """Process individual field values with type conversion and validation"""
        if pd.isna(value):
            return None if field_name not in ['control_ids', 'asset_ids', 'milestones'] else []
        
        # Convert to string and strip whitespace
        str_value = str(value).strip()
        
        if field_name in ['control_ids', 'asset_ids']:
            # Split comma-separated values
            if not str_value:
                return []
            return [item.strip() for item in str_value.split(',') if item.strip()]
        
        elif field_name == 'severity':
            # Validate severity values
            if str_value not in self.SEVERITY_VALUES:
                logger.warning(f"Invalid severity '{str_value}' at row {row_num}, column {excel_col}")
                # Try to map common variants
                str_value = self._normalize_severity(str_value)
            return str_value
        
        elif field_name == 'status':
            # Validate status values
            if str_value not in self.STATUS_VALUES:
                logger.warning(f"Invalid status '{str_value}' at row {row_num}, column {excel_col}")
                # Try to map common variants
                str_value = self._normalize_status(str_value)
            return str_value
        
        elif field_name in ['scheduled_completion_date', 'actual_completion_date']:
            # Handle date fields
            return self._parse_date(value, row_num, excel_col)
        
        else:
            return str_value if str_value else None
    
    def _normalize_severity(self, value: str) -> str:
        """Normalize severity values to standard format"""
        value_lower = value.lower()
        
        if value_lower in ['low', 'l']:
            return 'Low'
        elif value_lower in ['moderate', 'med', 'medium', 'm']:
            return 'Moderate'  
        elif value_lower in ['high', 'h']:
            return 'High'
        elif value_lower in ['critical', 'crit', 'c']:
            return 'Critical'
        
        return value  # Return original if no match
    
    def _normalize_status(self, value: str) -> str:
        """Normalize status values to standard format"""
        value_lower = value.lower()
        
        if value_lower in ['open', 'new']:
            return 'Open'
        elif value_lower in ['ongoing', 'in progress', 'in-progress']:
            return 'Ongoing'
        elif value_lower in ['completed', 'complete', 'closed', 'done']:
            return 'Completed'
        elif value_lower in ['risk accepted', 'accepted', 'risk_accepted']:
            return 'Risk Accepted'
        
        return value  # Return original if no match
    
    def _parse_date(self, value: Any, row_num: int, excel_col: str) -> Optional[str]:
        """Parse date values to ISO format"""
        if pd.isna(value):
            return None
        
        try:
            # Handle various date formats
            if isinstance(value, datetime):
                return value.strftime('%Y-%m-%d')
            elif isinstance(value, str):
                # Try to parse string dates
                date_obj = pd.to_datetime(value)
                return date_obj.strftime('%Y-%m-%d')
            else:
                # Try converting to datetime
                date_obj = pd.to_datetime(str(value))
                return date_obj.strftime('%Y-%m-%d')
        
        except (ValueError, TypeError):
            logger.warning(f"Invalid date '{value}' at row {row_num}, column {excel_col}")
            return str(value)  # Return original value as string
    
    def _extract_milestones(self, row: pd.Series, row_num: int) -> List[Dict[str, Any]]:
        """Extract milestone information from row"""
        milestones = []
        
        # Check if milestone columns exist
        milestone_desc = row.get('Milestone Description') or row.get('milestone_description')
        milestone_date = row.get('Milestone Date') or row.get('milestone_date')
        milestone_status = row.get('Milestone Status') or row.get('milestone_status')
        
        if milestone_desc and not pd.isna(milestone_desc):
            milestone = {
                'description': str(milestone_desc).strip(),
                'scheduled_date': self._parse_date(milestone_date, row_num, 'Milestone Date'),
                'status': str(milestone_status).strip() if not pd.isna(milestone_status) else 'Pending'
            }
            milestones.append(milestone)
        
        return milestones
    
    def _validate_row_data(self, row_data: Dict[str, Any], row_num: int) -> None:
        """Validate required fields in row data"""
        required_fields = ['poam_id', 'title', 'severity', 'status']
        
        for field in required_fields:
            if not row_data.get(field):
                raise ValueError(f"Missing required field '{field}' at row {row_num}")
    
    def _get_last_column_letter(self, num_columns: int) -> str:
        """Convert column number to Excel column letter"""
        result = ""
        while num_columns > 0:
            num_columns -= 1
            result = chr(num_columns % 26 + ord('A')) + result
            num_columns //= 26
        return result